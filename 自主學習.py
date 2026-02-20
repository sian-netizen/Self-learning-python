from openpyxl import Workbook , load_workbook
from datetime import datetime as time

status = 2
wb=Workbook()
i = []
m = []


def get_values(sheet):
    arr = []                      
    for row in sheet:
        arr2 = []                 
        for column in row:
            arr2.append(column.value)  
        arr.append(arr2)
    return arr




while status != 0:
    
    if status == 2:    #新增品項
        item = input('品項:')
        i.append(item)

        wb.create_sheet(item)
        ws = wb[item]
        money = int(input('售價:'))
        m.append(money)
        
        ws.append(['時間','數量','價格','獲利'])
        status = int(input('結束按0;新增購買紀錄按1;新增品項按2;切換工作表按3;查看工作表按4;刪除資料按4;更新資料按6:'))
    
    
    
    if status == 1:    #新增資料
        now = time.now()
        t = now.strftime('%H:%M:%S')
        number = int(input('數量'))   
        price = int(number)*money
        data=[
            {
                'time':t,
                'number':number,
                'money':money,
                'price':price,

            }

        ]
        print(data)
        for tea in data:
            ws.append(list(tea.values()))
        status = int(input('結束按0;新增購買紀錄按1;新增品項按2;切換工作表按3;查看工作表按4;刪除資料按4;更新資料按6:'))
        
        
    if status == 3:    #切換
        print(wb.sheetnames)
        sheet = input('想切換的工作表:')
        ws = wb[sheet]
        print(ws)
        money = m[i.index(sheet)]
        #money = ws['C2'].value
        status = int(input('結束按0;新增購買紀錄按1;新增品項按2;切換工作表按3;查看工作表按4;刪除資料按4;更新資料按6:'))
    
    if status == 4:    #查看
        print(wb.sheetnames)
        for value in get_values(wb[input('請輸入想查看的工作表:')]):
            print(value)
        status = int(input('結束按0;新增購買紀錄按1;新增品項按2;切換工作表按3;查看工作表按4;刪除資料按4;更新資料按6:'))    

    if status == 5:    #刪除
        print(wb.sheetnames)
        target_sheet = input('請輸入想刪除的資料所在工作表:')
        n = 1
        ws = wb[target_sheet]
        for value in get_values(ws):
            print(value, end = ' ' )
            print(n)
            n += 1
        target = int(input('請輸入刪除目標所在行列旁的數字:'))
        target_data = [cell.value for cell in ws[target]]
        print(f'已成功刪除資料{target_data}')
        ws.delete_rows(target)
        status = int(input('結束按0;新增購買紀錄按1;新增品項按2;切換工作表按3;查看工作表按4;刪除資料按4;更新資料按6:')) 
    
    if status == 6:   #更新
        print(wb.sheetnames)
        target_sheet = input('請輸入想取代的資料所在工作表:')
        n = 1
        ws = wb[target_sheet]
        for value in get_values(ws):
            print(value, end = ' ' )
            print(n)
            n += 1
        target = int(input('請輸入取代目標所在行列旁的數字:'))
        
        now = time.now()
        t = now.strftime('%H:%M:%S')
        money =  money = ws.cell(row = target, column = 3).value
        number = int(input('更改數量為:'))   
        price = int(number)*money
        new_data = [t, number, money, price]
        for cell, data in zip(ws[target], new_data):
            cell.value = data  
        
        print(new_data)
        status = int(input('結束按0;新增購買紀錄按1;新增品項按2;切換工作表按3;查看工作表按4;刪除資料按4;更新資料按6:')) 



wb.save('園遊會.xlsx')